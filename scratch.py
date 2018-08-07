#这是一个专用工具
#Version 1.2
#By aktomleader
#Start at 2018-08-04
#Python v3.6.0:41df79263a11
import openpyxl

# 加载工作簿
NA62 = openpyxl.load_workbook(r"D:\NA62.xlsx")
print("Load 100%")
sheetnames = NA62.sheetnames
print("表信息：\n",sheetnames)

# 获取sheet名称
Sheetname_Space = sheetnames[0]
Sheetname_Device = sheetnames[1]  #设备表
Sheetname_Dot = sheetnames[2]     #点位表
# 获取sheet
Sheet_Space  = NA62[Sheetname_Space]
Sheet_Device = NA62[Sheetname_Device]
Sheet_Dot    = NA62[Sheetname_Dot]

#设置对应列值
SheetDotC_column    = 3    # C列：3
SheetDeviceM_column = 13   # M列：13

#函数：取得点位所属设备的设备标识
def get_DeviceID(temp_SheetDotC_row):
    temp = str(Sheet_Dot.cell(temp_SheetDotC_row, SheetDotC_column).value)
    return temp

#函数：取得待对比设备-设备表-设备标识
def get_isDeviceID(temp_SheetDeviceM_row):
    temp = str(Sheet_Device.cell(temp_SheetDeviceM_row, SheetDeviceM_column).value)
    return temp

#函数：判断设备标识是否一致
def isThisDeviceID(temp_SheetDotC_row,temp_SheetDeviceM_row):
    temp_a = get_DeviceID(temp_SheetDotC_row)
    temp_b = get_isDeviceID(temp_SheetDeviceM_row)
    temp   = (temp_a == temp_b)
    return temp

#设置待处理行数
Sheet_Dot_C_maxcolumnNum    = 7+1
Sheet_Device_M_maxcolumnNum = Sheet_Device.max_row+1
#配置点位绝对路径
temp_SheetDotC_row   = 2
temp_SheetDeviceM_row = 2

while temp_SheetDotC_row < Sheet_Dot_C_maxcolumnNum:                                 #需要处理的行数：Sheet_Dot_C_maxcolumnNum 行
    print(temp_SheetDotC_row)
    temp_isThisDeviceID = isThisDeviceID(temp_SheetDotC_row,temp_SheetDeviceM_row)   #取得比较结果
    if temp_isThisDeviceID:                                                          #点位所属设备标识与待对比设备表设备标识一致，则取得对应设备表设备绝对路径与设备标识，合成点位绝对路径并存放在对应列
        temp_new_DeviceURL1 = str(Sheet_Device.cell(temp_SheetDeviceM_row,3).value)  #取得设备表设备绝对路径
        temp_new_DeviceURL2 = str(Sheet_Device.cell(temp_SheetDeviceM_row,11).value) #取得设备表设备标识
        temp_new_DeviceURL3 = '/'
        Sheet_Dot.cell(temp_SheetDotC_row,6).value = temp_new_DeviceURL1 + temp_new_DeviceURL2 + temp_new_DeviceURL3  #合成点位绝对路径并存放
        temp_SheetDotC_row    += 1     #进行下一行的处理
        temp_SheetDeviceM_row =  2     #初始化设备表待对比行
    elif temp_SheetDeviceM_row < Sheet_Device_M_maxcolumnNum:          #点位所属设备标识与待对比设备表设备标识不一致，则比对下一个设备标识
        temp_SheetDeviceM_row += 1
    else:                                                              # 点位表设备未在设备表找到对应设备，跳过
        temp_sheetDotC_row +=1
        Sheet_Dot.cell(temp_SheetDotC_row, 6).value = 'Device Unknow'

#点位映射字典
Dict_10301  = {'CT变比':'1001','平均线电压_V':'1002','零序电压_V':'1003','A_电流_A':'1004','B_电流_A':'1005'}
Dict_DeviceType = {'10301':Dict_10301}

#配置点位ID
temp_SheetDotC_row    = 2
temp_SheetDeviceM_row = 2

while temp_SheetDotC_row < Sheet_Dot_C_maxcolumnNum:                              #需要处理的行数：Sheet_Dot_C_maxcolumnNum 行
    print(temp_SheetDotC_row)
    temp_isThisDeviceID = isThisDeviceID(temp_SheetDotC_row, temp_SheetDeviceM_row)
    if temp_isThisDeviceID:                                                       #比对一致，则在字典中查找对应设备类型ID，并进一步查找点位映射的ID
        temp_DotName  = str(Sheet_Dot.cell(temp_SheetDotC_row,1).value)           #取得点位表点位名称
        temp_DeviceType   = str(Sheet_Device.cell(temp_SheetDeviceM_row,9).value) #取得设备表设备类型ID
        if temp_DotName in Dict_DeviceType[temp_DeviceType]:                      #此点位在字典中存在映射则查找字典取得点位ID并存放在对应列
            temp_DotID   = Dict_DeviceType[temp_DeviceType][temp_DotName]         #字典查询点位ID
            Sheet_Dot.cell(temp_SheetDotC_row,9).value = temp_DotID               #存放点位ID
            temp_SheetDotC_row    += 1                                            #进行下一个点位映射
            temp_SheetDeviceM_row =  2
        else:                                                                     #此点位在字典中不存在映射则写：Unknown
            Sheet_Dot.cell(temp_SheetDotC_row,9).value = 'Dot Unknown'
            temp_SheetDotC_row    += 1
            temp_SheetDeviceM_row =  2

    elif temp_SheetDeviceM_row < Sheet_Device_M_maxcolumnNum:                     #点位所属设备标识与待对比设备表设备标识不一致，则比对下一个设备标识
        temp_SheetDeviceM_row += 1

    else:                                                                         #点位表设备未在设备表找到对应设备，跳过
        temp_sheetDotC_row += 1
        Sheet_Dot.cell(temp_SheetDotC_row, 9).value = 'Device Unknow'

NA62.save("D:\\NA62.xlsx")
